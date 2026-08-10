# =====================================================================
# 
# INF2Calc Version 2026.08.05 (Alain DELIGNY - SGAR Normandie - Août 2026)
# 
# INF2Calc est une macro LibreOffice permettant l'ouverture dans Calc des restitutions Chorus INF
# exportées en Excel 2000 (faux fichier .xls, en réalité au format « Microsoft SpreadsheetML HTML »),
# après conversion automatique en fichier Excel (.xlsx) exploitable nativement par Calc.
# 
# Réalisé en Python, compatible avec l'interpréteur 3.11.14 (64 bit) intégré à LibreOffice 25.8.4.2 (X86_64)
# 
# - Licence :
#  Extension : Licence Ouverte / Open Licence Etalab v2.0
#     Présentation : https://www.etalab.gouv.fr/licence-ouverte-open-licence/
#     Texte complet (PDF) : https://www.etalab.gouv.fr/wp-content/uploads/2017/04/ETALAB-Licence-Ouverte-v2.0.pdf
#   Paternité : Alain DELIGNY - SGAR Normandie - Août 2026
#   Voir également l'article L. 121-1 du Code de la propriété intellectuelle
#   (droit au respect du nom de l'auteur) :
#     https://www.legifrance.gouv.fr/codes/article_lc/LEGIARTI000006278891
# - Composants inclus : openpyxl et et_xmlfile (Licence MIT)
# 
# Le détail complet des conditions figure dans la documentation
# 
# Historique & Contributeurs :
# 
# Version/Date : 2026.08.05
#   - Première version
#   - Alain DELIGNY (SGAR Normandie)
# 
# Arborescence installée :
#   
# Scripts/
# └── python/
#     ├── INF2Calc.py
#     └── pythonpath/
#         ├── openpyxl/
#         │   ├── __init__.py
#         │   └── ...
#         └── et_xmlfile/
#             ├── __init__.py
#             └── ...
# =====================================================================

import sys
import os

# Trouver pythonpath/ via sys.path pour les imports suivants
# (contourne le bug LO 25.8 qui n'ajoute pas pythonpath/ automatiquement)
# LO ajoute le dossier python/ à sys.path. On le repère et on remonte
# d'un niveau pour trouver le dossier pythonpath/ frère.
_script_dir = None
for p in sys.path:
    if os.path.basename(p) == 'python' and os.path.isdir(p):
        _script_dir = p
        break

if _script_dir:
    _pythonpath = os.path.join(os.path.dirname(_script_dir), 'pythonpath')
    _pythonpath = os.path.abspath(_pythonpath)
    if os.path.isdir(_pythonpath) and _pythonpath not in sys.path:
        sys.path.insert(0, _pythonpath)

import uno
import re
import html

from com.sun.star.awt import MessageBoxButtons, MessageBoxResults
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

def message_box(message, type = "MESSAGEBOX", buttons = MessageBoxButtons.BUTTONS_OK, title = ""):
    
    # MessageBoxType :
    # MESSAGEBOX   A normal message box. Without icon.
    # INFOBOX      A message box to inform the user about a certain event. With blue (i) icon. (BUTTONS_OK Only).
    # WARNINGBOX   A message to warn the user about a certain problem. With yellow /!\ icon.
    # ERRORBOX     A message box to provide an error message to the user. With red (X) icon.
    # QUERYBOX     A message box to query information from the user. With blue (?) icon.

    # MessageBoxButtons :
    # BUTTONS_OK = 1
    # BUTTONS_OK_CANCEL = 2
    # BUTTONS_YES_NO = 3
    # BUTTONS_YES_NO_CANCEL = 4
    # BUTTONS_RETRY_CANCEL = 5
    # BUTTONS_ABORT_IGNORE_RETRY = 6
    # DEFAULT_BUTTON_OK = 0x10000
    # DEFAULT_BUTTON_CANCEL = 0x20000
    # DEFAULT_BUTTON_RETRY = 0x30000
    # DEFAULT_BUTTON_YES = 0x40000
    # DEFAULT_BUTTON_NO = 0x50000
    # DEFAULT_BUTTON_IGNORE = 0x60000

    # MessageBoxResults :
    # CANCEL = 0
    # OK = 1
    # YES = 2
    # NO = 3
    # RETRY = 4
    # IGNORE = 5

    ctx = XSCRIPTCONTEXT.getComponentContext()
    smgr = ctx.getServiceManager()
    desktop = smgr.createInstanceWithContext("com.sun.star.frame.Desktop", ctx)
    doc = desktop.getCurrentComponent()
    
    try:
        parent_window = doc.getCurrentController().getFrame().getContainerWindow()
    except AttributeError:
        parent_window = desktop.getComponents().createEnumeration().nextElement().getCurrentController().getFrame().getContainerWindow()
        
    toolkit = smgr.createInstanceWithContext("com.sun.star.awt.Toolkit", ctx)
        
    msg_box = toolkit.createMessageBox(parent_window, uno.Enum("com.sun.star.awt.MessageBoxType", type), buttons, title, message)
    
    return msg_box.execute()

def get_file_path():

    ctx = XSCRIPTCONTEXT.getComponentContext()
    smgr = ctx.getServiceManager()

    file_picker = smgr.createInstanceWithContext("com.sun.star.ui.dialogs.FilePicker", ctx)

    file_picker.setTitle("Choisir le fichier INF Chorus")
    
    file_picker.appendFilter("INF Chorus (*.xls)", "*.xls")
    file_picker.appendFilter("Tous les fichiers (*.*)", "*.*")
    file_picker.setCurrentFilter("INF Chorus (*.xls)")

    # ExecutableDialogResults.OK == 1
    if file_picker.execute() == 1:
        selected_files = file_picker.getSelectedFiles()
        file_picker.dispose()
        if selected_files:
            return uno.fileUrlToSystemPath(selected_files[0])
        return None
    file_picker.dispose()
    return None

def is_ms_spreadsheetml_html(file_path):
    # Vérifie que le fichier est bien un export 'Page Web filtrée' / 'Microsoft SpreadsheetML HTML' d'Excel, et non :
    #   - un vrai fichier binaire .xls (BIFF / conteneur OLE2), signature D0 CF 11 E0
    #   - un export SpreadsheetML XML (racine <?xml ... <Workbook ...>), format différent bien que parfois confondu avec le nôtre.
    # La signature retenue est la présence simultanée, en tête de fichier, de :
    #   - une balise <html ...>
    #   - le namespace 'urn:schemas-microsoft-com:office:excel'
    #   - le meta ProgId 'Excel.Sheet'
   
    try:
        with open(file_path, "rb") as f:
            header = f.read(4096)
    
        # Vrai binaire .xls (OLE2/BIFF) : signature D0 CF 11 E0 A1 B1 1A E1
        if header[:4] == b"\xD0\xCF\x11\xE0":
            return False

        try:
            text = header.decode("utf-8", errors="ignore")
        except UnicodeDecodeError:
            text = header.decode("cp1252", errors="ignore")

        text_lower = text.lower()

        return (
            "<html" in text_lower 
            and "urn:schemas-microsoft-com:office:excel" in text_lower 
            and "progid" in text_lower and "excel.sheet" in text_lower
        )
    
    except OSError:
        return False


def ms_spreadsheetml_html_to_xlsx(source_file, xlsx_file):

    with open(source_file, "rb") as f:
        raw = f.read()

    # --- Encodage : BOM, sinon meta charset, sinon repli utf-8/cp1252/latin-1 ---
    boms = ((b"\xEF\xBB\xBF", "utf-8-sig"), (b"\xFF\xFE", "utf-16-le"), (b"\xFE\xFF", "utf-16-be"))
    encoding = next((enc for bom, enc in boms if raw.startswith(bom)), None)
    if not encoding:
        m = re.search(rb'charset=["\']?([\w-]+)', raw[:8192], re.IGNORECASE)
        candidats = ([m.group(1).decode()] if m else []) + ["utf-8", "windows-1252", "iso-8859-1"]
        encoding = min(candidats, key=lambda enc: raw.decode(enc, errors="replace").count("\ufffd"))
    content = raw.decode(encoding, errors="replace")

    # --- Classes de style CSS (.xN {...}) ---
    css = {}
    for cls, body in re.findall(r"\.(x\d+)\s*\{([^}]*)\}", content):
        props = {}
        for decl in body.split(";"):
            if ":" in decl:
                k, v = decl.split(":", 1)
                props[k.strip().lower()] = v.strip()
        css[cls] = props

    def build_style(classname):
        # Traduit une classe CSS ('x9', 'x2', ...) en styles openpyxl.
        # Appelée une fois par cellule stylée : c'est la seule fonction
        # imbriquée réellement réutilisée, donc la seule qui se justifie.
        props = css.get(classname, {})
        font_kwargs = {}
        if "color" in props:
            font_kwargs["color"] = props["color"].lstrip("#").upper()
        if "font-size" in props:
            try:
                font_kwargs["size"] = float(props["font-size"].replace("pt", ""))
            except ValueError:
                pass
        if "font-family" in props:
            font_kwargs["name"] = props["font-family"].strip("\"' ")
        if props.get("font-weight") in ("700", "bold"):
            font_kwargs["bold"] = True

        fill = None
        bg = props.get("background", "")
        if re.match(r"#[0-9a-fA-F]{6}", bg):
            fill = PatternFill(fill_type="solid", fgColor=bg.lstrip("#").upper())

        def side(edge):
            m = re.match(r"([\d.]+)pt\s+\w+\s+(#[0-9a-fA-F]{6})", props.get(f"border-{edge}", ""))
            if not m:
                return Side()
            pt, color = m.groups()
            style = "thin" if float(pt) <= 1 else "medium" if float(pt) <= 2.25 else "thick"
            return Side(style=style, color=color.lstrip("#").upper())

        border = Border(left=side("left"), right=side("right"), top=side("top"), bottom=side("bottom"))
        align = Alignment(
            horizontal={"left": "left", "right": "right", "center": "center", "justify": "justify"}.get(props.get("text-align")),
            vertical={"top": "top", "middle": "center", "bottom": "bottom"}.get(props.get("vertical-align")),
            wrap_text=True,
        )
        return Font(**font_kwargs), fill, border, align

    # --- Onglets déclarés / tables HTML ---
    # Chorus n'insère aucun séparateur exploitable entre plusieurs <table>
    # d'un même onglet : les tables excédentaires sont donc empilées à la
    # suite du dernier onglet déclaré, dans l'ordre du document.
    sheet_names = re.findall(r"<x:ExcelWorksheet>\s*<x:Name>([^<]*)</x:Name>", content) or ["Feuille1"]
    tables = re.findall(r"<table[^>]*>(.*?)</table>", content, re.S)
    cell_re = re.compile(r"<t[dh]([^>]*)>(.*?)</t[dh]>", re.S)
    attr_re = re.compile(r'(\w[\w:-]*)\s*=\s*"([^"]*)"')

    classeur = Workbook()
    classeur.remove(classeur.active)
    feuille, idx_feuille, ligne_suivante, fusionnees = None, None, 1, {}

    for i, table_html in enumerate(tables):
        idx = min(i, len(sheet_names) - 1)
        if idx != idx_feuille:
            nom = re.sub(r'[:\\/?*\[\]]', "_", sheet_names[idx])[:31] or f"Feuille{idx + 1}"
            feuille, idx_feuille, ligne_suivante, fusionnees = classeur.create_sheet(title=nom), idx, 1, {}

        lignes_html = re.findall(r"<tr[^>]*>(.*?)</tr>", table_html, re.S)
        derniere_ligne = ligne_suivante - 1
        for decalage, ligne_html in enumerate(lignes_html):
            r = ligne_suivante + decalage
            c = 1
            for attrs_brut, contenu_cellule in cell_re.findall(ligne_html):
                attrs = dict(attr_re.findall(attrs_brut))
                while fusionnees.get((r, c)):
                    c += 1
                colspan, rowspan = int(attrs.get("colspan", 1)), int(attrs.get("rowspan", 1))
                texte = html.unescape(re.sub(r"<[^>]+>", "", contenu_cellule)).replace("\xa0", " ").strip()

                valeur, format_nombre = None, None
                num = attrs.get("x:num")
                if num is not None:
                    # x:num porte la valeur numérique exacte fournie par Excel
                    try:
                        valeur = float(num)
                    except ValueError:
                        valeur = texte
                    m = re.search(r"mso-number-format:'([^']*)'", attrs.get("style", ""))
                    if m:
                        format_nombre = m.group(1).split(";")[0]
                elif texte:
                    valeur = texte

                cellule = feuille.cell(row=r, column=c, value=valeur)
                if format_nombre and format_nombre != "General":
                    cellule.number_format = format_nombre
                if attrs.get("class"):
                    police, remplissage, bordure, alignement = build_style(attrs["class"])
                    cellule.font, cellule.border, cellule.alignment = police, bordure, alignement
                    if remplissage:
                        cellule.fill = remplissage

                if colspan > 1 or rowspan > 1:
                    feuille.merge_cells(start_row=r, start_column=c, end_row=r + rowspan - 1, end_column=c + colspan - 1)
                    for rr in range(r, r + rowspan):
                        for cc in range(c, c + colspan):
                            fusionnees[(rr, cc)] = True
                c += colspan
            derniere_ligne = max(derniere_ligne, r)
        ligne_suivante = derniere_ligne + 1

    # =====================================================================
    # Post-traitement : police unique, suppression du renvoi à la ligne,
    # largeurs de colonnes optimales (hors ligne 1 de titre)
    # =====================================================================

    for feuille in classeur.worksheets:

        # Cellules appartenant à une fusion MULTI-COLONNES : leur contenu
        # (typiquement le titre fusionné) ne doit pas gonfler la largeur
        # d'une colonne isolée.
        fusion_multi = set()
        for plage in feuille.merged_cells.ranges:
            if plage.max_col > plage.min_col:
                for ligne in feuille[plage.coord]:
                    for cel in ligne:
                        fusion_multi.add(cel.coordinate)

        largeurs = {}   # n° de colonne -> longueur max observée

        for ligne in feuille.iter_rows():
            for cel in ligne:

                # 1. Police "Liberation Sans" partout, en conservant les
                #    autres attributs posés par build_style (taille, gras,
                #    couleur...). Les objets de style openpyxl étant
                #    partagés entre cellules, on REconstruit un Font au
                #    lieu de modifier l'existant.
                p = cel.font
                cel.font = Font(name="Liberation Sans", size=p.size,
                                bold=p.bold, italic=p.italic,
                                underline=p.underline, color=p.color)

                # 2. Suppression du "Renvoi à la ligne automatique"
                #    (wrap_text), en conservant les alignements.
                a = cel.alignment
                cel.alignment = Alignment(horizontal=a.horizontal,
                                          vertical=a.vertical,
                                          wrap_text=False)

                # 3. Collecte des longueurs pour la largeur optimale :
                #    - ligne 1 (titre) ignorée,
                #    - cellules de fusions multi-colonnes ignorées,
                #    - cellules fantômes des fusions (MergedCell) ignorées.
                if cel.row == 1 or isinstance(cel, MergedCell) \
                        or cel.coordinate in fusion_multi:
                    continue
                v = cel.value
                if v is None:
                    continue
                # Approximation de la longueur AFFICHÉE : les montants
                # Chorus sont rendus avec séparateurs de milliers et
                # 2 décimales, plus larges que leur repr Python.
                s = f"{v:,.2f}" if isinstance(v, float) else str(v)
                # En cas de contenu multi-lignes, seule la plus longue
                # ligne compte pour la largeur.
                longueur = max(len(part) for part in s.split("\n"))
                if longueur > largeurs.get(cel.column, 0):
                    largeurs[cel.column] = longueur

        # Application des largeurs : l'unité openpyxl vaut ~1 caractère de
        # la police par défaut ; +10 % et +2 caractères de marge, avec un
        # plancher (colonnes vides lisibles) et un plafond de sécurité.
        for col, longueur in largeurs.items():
            feuille.column_dimensions[get_column_letter(col)].width = \
                min(max(longueur * 1.1 + 2, 8), 60)

    classeur.save(xlsx_file)

def INF2Calc(*args):
    
    source_file = get_file_path()
    if not source_file:
        return
    
    if not is_ms_spreadsheetml_html(source_file):
        message_box(
            f"Ce fichier n'est pas un export Excel 2000 d'une INF Chorus\nou un fichier Microsoft SpreadsheetML HTML.\n\nFichier : '{source_file}'\n ", 
            "ERRORBOX", 
            MessageBoxButtons.BUTTONS_OK, 
            "Format non reconnu"
        )
        return

    if message_box(
            f"Ce fichier est un fichier export Excel 2000 d'une INF Chorus\n(Fichier Microsoft SpreadsheetML HTML).\n\nFichier : '{source_file}'\n\nVoulez vous le convertir en fichier Excel (.xlsx) ?\n ", 
            "QUERYBOX", 
            MessageBoxButtons.BUTTONS_YES_NO, 
            "Conversion de fichier Chorus INF"
        ) != MessageBoxResults.YES:
        return

    path = os.path.dirname(source_file)
    file_name, _ext = os.path.splitext(os.path.basename(source_file))
    xlsx_file = os.path.join(path, f"{file_name}.xlsx")

    ms_spreadsheetml_html_to_xlsx(source_file, xlsx_file)

    # Ouverture du fichier généré, selon la même logique que Fichier/Ouvrir :
    # "_default" réutilise la fenêtre courante si elle est vide (document
    # vierge non modifié), sinon ouvre une nouvelle fenêtre - c'est le
    # comportement documenté du mot-clé réservé "_default" de
    # XComponentLoader, spécifiquement prévu pour ça.
    url_xlsx = uno.systemPathToFileUrl(xlsx_file)
    ctx = XSCRIPTCONTEXT.getComponentContext()
    smgr = ctx.getServiceManager()
    desktop = smgr.createInstanceWithContext("com.sun.star.frame.Desktop", ctx)

    try:
        desktop.loadComponentFromURL(url_xlsx, "_default", 0, ())
    except Exception as e:
        # Sans ce filet, une erreur d'ouverture (fichier verrouillé, xlsx
        # corrompu, etc.) remonterait comme une trace UNO brute au lieu
        # d'un message compréhensible, contrairement au Fichier/Ouvrir natif.
        message_box(
            f"Le fichier converti n'a pas pu être ouvert :\n{xlsx_file}\n\n{e}",
            "ERRORBOX",
            MessageBoxButtons.BUTTONS_OK,
            "Erreur à l'ouverture"
        )

g_exportedScripts = (INF2Calc,)
